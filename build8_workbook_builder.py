"""Excel workbook builder (build 8).

Reads assignF.json, hoursF.json, training.json produced by gen7_engine.py
and builds the 11-tab schedule workbook: Schedule, seven per-day coverage
timelines, Coverage Check, Hours Summary, Notes & Decisions.

Run from the repo root, after the engine:
    python3 gen7_engine.py
    python3 build8_workbook_builder.py
Output: sample_output/Crew_Schedule_SAMPLE.xlsx
"""
import json
from collections import defaultdict
from schedule_data import *  # DEMAND, CREW, fmt, DAYS, POS_ORDER, CT, KIT/DT/FC
# (originally exec(open('gen7.py')...) pulled the engine's data section in; the shared
#  data now lives in schedule_data.py and both scripts import it - see README)
POS_ORDER[:]=["Board1","Grill","DTExp","Board2","Board3","Toast","DTOrd","DTOI","Cash"]
for _d in ("Thu","Fri","Sat","Sun"):
    _w=DEMAND[_d].get("Board2",[])
    if _w:
        _i=min(range(len(_w)),key=lambda j:_w[j][0])
        if _w[_i][0]>9: _w[_i]=(9.0,_w[_i][1])
        DEMAND[_d]["Board2"]=_w
assign=json.load(open("assignF.json")); hoursF=json.load(open("hoursF.json"))
TRAIN=json.load(open("training.json"))
TRAIN_SET={(p,d,pos,round(s,3),round(e,3)) for (p,d,pos,s,e) in TRAIN}
def is_train(n,d,pos,s,e): return (n,d,pos,round(s,3),round(e,3)) in TRAIN_SET

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# display names shown in the workbook - pseudonyms, real roster withheld (see README)
NAMES={"Avery":"Avery Sutton","Quinn":"Quinn Delgado","Tessa":"Tessa Vance","Emil":"Emil Navarro",
"Priya":"Priya Chandra","Miles":"Miles Ferreira","Lena":"Lena Whitfield","Wren":"Wren Okafor",
"Theo":"Theo Marsh","Bianca":"Bianca Rowe","Dara":"Dara Kimura","Felix":"Felix Arroyo",
"Harper":"Harper Nunes","Ivy":"Ivy Sutton","Paloma":"Paloma Reyes","Kofi":"Kofi Mensah",
"Silas":"Silas Duarte","Hugo":"Hugo Beltran","Kwame":"Kwame Boateng","Mateo":"Mateo Fuentes",
"Peyton":"Peyton Calloway","Beckett":"Beckett Hale","Rowan":"Rowan Pierce","Deshawn":"Deshawn Pratt",
"Selena":"Selena Hart","Marco":"Marco Renner","Dante":"Dante Moreau","Beau":"Beau Hendrix",
"Lionel":"Lionel Ford","Ansel":"Ansel Gray","Amina":"Amina Sow","Marlowe":"Marlowe Bishop",
"Idris":"Idris Sharif","Rosa":"Rosa Delacruz","Porter":"Porter Lindqvist","JaNae":"Ja'Nae Winslow",
"Celeste":"Celeste Byrne","Bishal":"Bishal Thapa","Estelle":"Estelle Fontaine","Juniper":"Juniper Vale",
"Genevieve":"Genevieve Stroud","Ezra":"Ezra Lockhart",
"Camila":"Camila Sutton","Astrid":"Astrid Larkin","Odette":"Odette Marchetti",
"Simone":"Simone Deveraux","Freya":"Freya Adjei"}
DISP={"Grill":"Grill","Board1":"Board 1","Board2":"Board 2","Board3":"Board 3","Toast":"Toast","DTExp":"DT Expo","DTOrd":"DT Order","DTOI":"DT Out","Cash":"Cashier","RI":"RI"}
STN={"Grill":"K","Board1":"K","Board2":"K","Board3":"K","Toast":"K","DTExp":"D","DTOrd":"D","DTOI":"D","Cash":"F","RI":"I"}
DATES={"Wed":"Wed 7/1","Thu":"Thu 7/2","Fri":"Fri 7/3","Sat":"Sat 7/4","Sun":"Sun 7/5","Mon":"Mon 7/6","Tue":"Tue 7/7"}
GROUPS=[("PERFORMERS \u2014 priority crew, scheduled first",
         ["Quinn","Avery","Emil","Tessa","Dara","Theo","Harper","Hugo","Kwame","Rosa",
          "Paloma","Rowan","Ivy","Ezra","Bianca","Beckett","Peyton","Amina","Marco","Felix",
          "Ansel","Juniper","Genevieve","Marlowe","Kofi","Silas"]),
        ("FILL-INS \u2014 used only where performers couldn't cover",
         ["Priya","Lena","Wren","Mateo","Selena","Dante","Lionel","Idris","JaNae","Celeste","Bishal","Odette","Simone","Freya"]),
        ("TRAINEES \u2014 shadow a CT on station (extra body, NOT coverage)",
         ["Camila","Astrid"]),
        ("RESTAURANT IMAGING (RI) + CIP",
         ["Miles"])]

ARIAL=lambda **k: Font(name="Arial",**k)
thin=Side(style="thin",color="D0D0D0"); BORDER=Border(left=thin,right=thin,top=thin,bottom=thin)
FILL={"K":PatternFill("solid",fgColor="FFF3E0"),"D":PatternFill("solid",fgColor="E3F0FB"),
      "F":PatternFill("solid",fgColor="E8F5E9"),"I":PatternFill("solid",fgColor="ECECEC")}
HEADFILL=PatternFill("solid",fgColor="1F3864"); GRPFILL=PatternFill("solid",fgColor="2E5496")
REDFILL=PatternFill("solid",fgColor="F4B7B7"); GREENFILL=PatternFill("solid",fgColor="C8E6C9")
YEL=PatternFill("solid",fgColor="FFF2CC")
TRAINFILL=PatternFill("solid",fgColor="C9A0DC")  # purple = trainee shadowing a CT
GRPFILL2=PatternFill("solid",fgColor="8EA9DB")
def dash(s,e): return f"{fmt(s)}\u2013{fmt(e)}"

# per person per day shift strings
pday=defaultdict(lambda:defaultdict(list))
for n,lst in assign.items():
    for (d,pos,s,e) in lst: pday[n][d].append((s,e,pos))
for n in pday:
    for d in pday[n]:
        pday[n][d].sort()
        merged=[]
        for (s_,e_,p_) in pday[n][d]:
            if merged and merged[-1][2]==p_ and abs(merged[-1][1]-s_)<1e-6:
                merged[-1]=(merged[-1][0],e_,p_)
            else: merged.append((s_,e_,p_))
        pday[n][d]=merged

wb=Workbook()
# ============ SHEET 1: SCHEDULE ============
ws=wb.active; ws.title="Schedule"
ws["A1"]="CREW SCHEDULE (ANONYMIZED SAMPLE)   |   Week of Wed 7/1 - Tue 7/7/26"
ws["A1"].font=ARIAL(bold=True,size=14,color="FFFFFF"); ws["A1"].fill=HEADFILL
ws["A1"].alignment=Alignment(vertical="center",horizontal="left",indent=1)
ws.merge_cells("A1:J1"); ws.row_dimensions[1].height=26
ws["A2"]="Stations:  K = Kitchen (Grill/Board/Toast)   ·   D = Drive Thru (Expo/Order/In-Out)   ·   F = Front Counter (Cashier)   ·   I = Restaurant Imaging (RI)"
ws["A2"].font=ARIAL(italic=True,size=9,color="555555"); ws.merge_cells("A2:J2")
hdr=["Crew","Tier"]+[DATES[d] for d in DAYS]+["Wk Hrs"]
r=3
for c,h in enumerate(hdr,1):
    cell=ws.cell(r,c,h); cell.font=ARIAL(bold=True,size=10,color="FFFFFF"); cell.fill=GRPFILL
    cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); cell.border=BORDER
ws.row_dimensions[r].height=22
r=4
_TIERMAP=["Performer","Fill-in","Trainee","RI"]
TIERLBL={n:_TIERMAP[gi] for gi,(t,mem) in enumerate(GROUPS) for n in mem}
for title,members in GROUPS:
    ws.cell(r,1,title); ws.cell(r,1).font=ARIAL(bold=True,size=10,color="FFFFFF")
    for c in range(1,11): ws.cell(r,c).fill=GRPFILL; ws.cell(r,c).border=BORDER
    ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=10); r+=1
    members=sorted(members,key=lambda n:-hoursF.get(n,0))
    for n in members:
        ws.cell(r,1,NAMES[n]).font=ARIAL(bold=True,size=10)
        ws.cell(r,2,TIERLBL.get(n,"Fill")).font=ARIAL(size=8,color="666666")
        for ci,d in enumerate(DAYS,3):
            shifts=pday[n].get(d,[])
            if shifts:
                parts=[]; alltrain=True
                for (s,e,p) in shifts:
                    if is_train(n,d,p,s,e): parts.append(f"TRAIN {DISP[p]} {dash(s,e)}")
                    elif p=="RI": parts.append(f"{DISP[p]} {dash(s,e)}"+("  +CIP" if d in ("Thu","Tue") else "")); alltrain=False
                    else: parts.append(f"{DISP[p]} {dash(s,e)}"); alltrain=False
                txt="\n".join(parts)
                cell=ws.cell(r,ci,txt); st=STN[shifts[0][2]]
                cell.fill=TRAINFILL if alltrain else FILL[st]; cell.font=ARIAL(size=8)
            else:
                cell=ws.cell(r,ci,""); 
            cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); cell.border=BORDER
        hc=ws.cell(r,10,round(hoursF.get(n,0),1)); hc.font=ARIAL(bold=True,size=10)
        hc.alignment=Alignment(horizontal="center",vertical="center"); hc.border=BORDER
        for c in (1,2): ws.cell(r,c).border=BORDER; ws.cell(r,c).alignment=Alignment(vertical="center",wrap_text=True)
        ws.row_dimensions[r].height=30; r+=1
ws.column_dimensions["A"].width=20; ws.column_dimensions["B"].width=11
for d in range(3,10): ws.column_dimensions[get_column_letter(d)].width=15
ws.column_dimensions["J"].width=8
ws.freeze_panes="C4"

# ============ SHEET 2: COVERAGE CHECK ============
cv=wb.create_sheet("Coverage Check")
cv["A1"]="COVERAGE CHECK - bodies on each position per hour vs. deployment requirement"
cv["A1"].font=ARIAL(bold=True,size=13,color="FFFFFF"); cv["A1"].fill=HEADFILL
HRS=list(range(8,25))  # 8a .. 12a(midnight start) ; bucket h = h:00-h+1:00 ; 24 = 12a-1a
cv.merge_cells(start_row=1,start_column=1,end_row=1,end_column=2+len(HRS))
cv["A2"]="Number = crew scheduled on that position that hour.  RED = required but uncovered (gap).  Green = covered.  Blank = not required."
cv["A2"].font=ARIAL(italic=True,size=9,color="555555"); cv.merge_cells(start_row=2,start_column=1,end_row=2,end_column=2+len(HRS))
def req_at(d,pos,h):
    return sum(1 for (s,e) in DEMAND[d].get(pos,[]) if s<=h+0.5<e)
def asg_at(d,pos,h):
    return sum(1 for nn in assign for (dd,pp,s,e) in assign[nn] if dd==d and pp==pos and not is_train(nn,d,pos,s,e) and s<=h+0.5<e)
row=3
for d in DAYS:
    cv.cell(row,1,DATES[d]).font=ARIAL(bold=True,size=11,color="FFFFFF")
    for c in range(1,3+len(HRS)): cv.cell(row,c).fill=GRPFILL
    cv.merge_cells(start_row=row,start_column=1,end_row=row,end_column=2+len(HRS)); row+=1
    cv.cell(row,1,"Position").font=ARIAL(bold=True,size=9); cv.cell(row,1).border=BORDER
    cv.cell(row,2,"Req").font=ARIAL(bold=True,size=9); cv.cell(row,2).border=BORDER
    for i,h in enumerate(HRS):
        c=cv.cell(row,3+i,fmt(h).replace(":00","")); c.font=ARIAL(bold=True,size=8)
        c.alignment=Alignment(horizontal="center"); c.border=BORDER
    row+=1
    for pos in POS_ORDER:
        if not DEMAND[d].get(pos): continue
        cv.cell(row,1,DISP[pos]).font=ARIAL(size=9); cv.cell(row,1).fill=FILL[STN[pos]]; cv.cell(row,1).border=BORDER
        peakreq=max(req_at(d,pos,h) for h in HRS)
        cv.cell(row,2,peakreq).font=ARIAL(size=8,color="888888"); cv.cell(row,2).alignment=Alignment(horizontal="center"); cv.cell(row,2).border=BORDER
        for i,h in enumerate(HRS):
            rq=req_at(d,pos,h); ag=asg_at(d,pos,h); c=cv.cell(row,3+i)
            c.alignment=Alignment(horizontal="center"); c.border=BORDER; c.font=ARIAL(size=8)
            if rq==0: c.value=""
            elif ag<rq: c.value=ag; c.fill=REDFILL; c.font=ARIAL(size=8,bold=True,color="A00000")
            else: c.value=ag; c.fill=GREENFILL
        row+=1
    row+=1
cv.column_dimensions["A"].width=12; cv.column_dimensions["B"].width=5
for i in range(len(HRS)): cv.column_dimensions[get_column_letter(3+i)].width=4.5
cv.freeze_panes="C3"

# ============ SHEET 3: HOURS SUMMARY ============
hs=wb.create_sheet("Hours Summary")
hs["A1"]="HOURS SUMMARY - weekly hours vs. tier target"
hs["A1"].font=ARIAL(bold=True,size=13,color="FFFFFF"); hs["A1"].fill=HEADFILL
hs.merge_cells("A1:M1")
hh=["Crew","Tier"]+[DATES[d] for d in DAYS]+["Week","Target","+/-","Status"]
for c,h in enumerate(hh,1):
    cell=hs.cell(2,c,h); cell.font=ARIAL(bold=True,size=10,color="FFFFFF"); cell.fill=GRPFILL
    cell.alignment=Alignment(horizontal="center",wrap_text=True); cell.border=BORDER
hs.row_dimensions[2].height=22
rr=3
for title,members in GROUPS:
    for n in sorted(members,key=lambda n:-hoursF.get(n,0)):
        hs.cell(rr,1,NAMES[n]).font=ARIAL(size=10,bold=True); hs.cell(rr,1).border=BORDER
        hs.cell(rr,2,TIERLBL.get(n,"Fill")).font=ARIAL(size=8,color="666666"); hs.cell(rr,2).border=BORDER
        for ci,d in enumerate(DAYS,3):
            v=sum(e-s for (s,e,p) in pday[n].get(d,[]))
            c=hs.cell(rr,ci, round(v,1) if v else None); c.alignment=Alignment(horizontal="center"); c.border=BORDER; c.font=ARIAL(size=9)
        wk=get_column_letter(3); en=get_column_letter(9)
        hs.cell(rr,10,f"=SUM({wk}{rr}:{en}{rr})").font=ARIAL(bold=True,size=10)
        hs.cell(rr,10).alignment=Alignment(horizontal="center"); hs.cell(rr,10).border=BORDER
        hs.cell(rr,11,CREW[n]["target"]).font=ARIAL(size=10); hs.cell(rr,11).alignment=Alignment(horizontal="center"); hs.cell(rr,11).border=BORDER
        hs.cell(rr,12,f"=J{rr}-K{rr}").font=ARIAL(size=10); hs.cell(rr,12).alignment=Alignment(horizontal="center"); hs.cell(rr,12).border=BORDER
        hs.cell(rr,13,f'=IF(J{rr}<K{rr}-3.5,"under",IF(J{rr}>K{rr}+3.5,"over","on target"))').font=ARIAL(size=9)
        hs.cell(rr,13).alignment=Alignment(horizontal="center"); hs.cell(rr,13).border=BORDER
        rr+=1
hs.cell(rr,1,"TOTAL").font=ARIAL(bold=True,size=10); hs.cell(rr,10,f"=SUM(J3:J{rr-1})").font=ARIAL(bold=True,size=10)
hs.cell(rr,10).alignment=Alignment(horizontal="center")
hs.cell(rr,11,"deployment need:"); hs.cell(rr,11).font=ARIAL(size=9,italic=True)
hs.cell(rr,12,750).font=ARIAL(size=10,bold=True)  # representative - real weekly deployment total redacted

# labor budget per day vs deployment optimal/target
# representative targets - real values redacted
# (the store's actual Optimal/Target labor hours come from the company's internal
#  deployment forecasts; these are round placeholder numbers of similar magnitude)
OPT={"Wed":None,"Thu":None,"Fri":120,"Sat":125,"Sun":125,"Mon":90,"Tue":90}
TGT={"Wed":None,"Thu":None,"Fri":100,"Sat":105,"Sun":105,"Mon":75,"Tue":75}
br=rr+3
hs.cell(br,1,"LABOR BUDGET CHECK - scheduled hrs vs. deployment").font=ARIAL(bold=True,size=11,color="FFFFFF")
for c in range(1,6): hs.cell(br,c).fill=GRPFILL
hs.merge_cells(start_row=br,start_column=1,end_row=br,end_column=5); br+=1
for c,h in enumerate(["Day","Scheduled","Optimal","Target","vs Optimal"],1):
    cc=hs.cell(br,c,h); cc.font=ARIAL(bold=True,size=9); cc.fill=GRPFILL2; cc.border=BORDER; cc.alignment=Alignment(horizontal="center")
br+=1
for d in DAYS:
    sched=round(sum(sum(e-s for (s,e,p) in pday[n].get(d,[])) for n in pday),1)
    hs.cell(br,1,DATES[d]).font=ARIAL(size=9,bold=True); hs.cell(br,1).border=BORDER
    hs.cell(br,2,sched).font=ARIAL(size=9); hs.cell(br,2).alignment=Alignment(horizontal="center"); hs.cell(br,2).border=BORDER
    hs.cell(br,3,OPT[d] if OPT[d] else "-").font=ARIAL(size=9); hs.cell(br,3).alignment=Alignment(horizontal="center"); hs.cell(br,3).border=BORDER
    hs.cell(br,4,TGT[d] if TGT[d] else "-").font=ARIAL(size=9); hs.cell(br,4).alignment=Alignment(horizontal="center"); hs.cell(br,4).border=BORDER
    if OPT[d]:
        diff=round(sched-OPT[d],1); c=hs.cell(br,5,diff); c.alignment=Alignment(horizontal="center"); c.border=BORDER
        c.font=ARIAL(size=9,bold=True,color=("A00000" if diff>3 else "1F7A1F"))
    else:
        hs.cell(br,5,"-").font=ARIAL(size=9); hs.cell(br,5).alignment=Alignment(horizontal="center"); hs.cell(br,5).border=BORDER
    br+=1
hs.cell(br,1,"Optimal/Target are representative placeholder values (the real deployment numbers are redacted for publication). Negative = under budget (good).").font=ARIAL(italic=True,size=8,color="666666")
hs.merge_cells(start_row=br,start_column=1,end_row=br,end_column=5)
hs.column_dimensions["A"].width=20; hs.column_dimensions["B"].width=12
for d in range(3,10): hs.column_dimensions[get_column_letter(d)].width=8
for cc in ("J","K","L","M"): hs.column_dimensions[cc].width=9
hs.column_dimensions["M"].width=11; hs.freeze_panes="C3"

# ============ SHEET 4: NOTES & DECISIONS ============
nt=wb.create_sheet("Notes & Decisions")
nt.column_dimensions["A"].width=120
def line(txt,bold=False,fill=None,size=10,color="000000"):
    global nrow
    c=nt.cell(nrow,1,txt); c.font=ARIAL(bold=bold,size=size,color=color); c.alignment=Alignment(wrap_text=True,vertical="top")
    if fill: c.fill=fill
    nt.row_dimensions[nrow].height=15*(1+txt.count("\n")+len(txt)//110); nrow+=1
nrow=1
line("SCHEDULE NOTES & KEY DECISIONS   (Week of Wed 7/1 \u2013 Tue 7/7/26)",bold=True,size=13,fill=HEADFILL,color="FFFFFF")
line("")
line("THIS IS A PERFORMANCE-PRIORITY SCHEDULE",bold=True,size=11,fill=GRPFILL,color="FFFFFF")
line("\u2022 Your 27 PERFORMERS are scheduled FIRST and loaded to cover the week. The rest of the crew are FILL-INS, used only where a performer couldn't cover the slot. Result this week: performers carry ~667 hrs, fill-ins ~104 hrs.")
line("\u2022 PERFORMERS: Quinn, Avery, Emil, Tessa, Dara, Theo, Harper Nunes, Hugo, Kwame, Rosa, Paloma, Rowan, Ivy, Ezra, Bianca, Beckett, Peyton Calloway, Amina, Marco, Felix, Ansel, Juniper Vale, Genevieve Stroud, Marlowe, Kofi, Silas, and Miles (on RI).")
line("\u2022 FILL-INS used this week (only as needed): Priya, Lena, Selena, Lionel, Idris, Dante, Wren, Mateo, Ja'Nae, Celeste, Bishal. NOT needed this week: Odette, Simone, Freya (lowest-priority fill-ins \u2014 0 hrs).")
line("")
line("HEADS-UP \u2014 PERFORMERS RUNNING OVER THEIR USUAL HOURS",bold=True,size=11,fill=GRPFILL,color="FFFFFF")
line("\u2022 Your performers' combined usual target hours (~535) are LESS than the week's demand (~748). To cover the week with performers first, several run above their normal target: e.g. Dara +11, Theo +10, Hugo +10, Kwame +10, Ansel +10, Beckett +11, Ivy +12. I capped the overage at +12 so nobody is pushed to a crazy number.",fill=YEL)
line("\u2022 If some of those performers can't actually take that many hours, tell me each person's real weekly max and I'll let a few more fill-in shifts absorb the difference. If you'd rather push fill-ins even lower, I can raise the cap and a few performers will run closer to 40.",fill=YEL)
line("\u2022 A few performers land UNDER their target (Tessa, Bianca, Felix) \u2014 that's their own availability limiting them, not deprioritization. They got every slot they were available for.")
line("")
line("RESTAURANT IMAGING (RI) \u2014 MILES, 8:00a\u201312:00p  (built off the June standards audit)",bold=True,size=11,fill=GRPFILL,color="FFFFFF")
line("\u2022 Miles runs RI 8:00a\u201312:00p on Wed/Thu/Fri/Mon/Tue (off Sat/Sun); CIP (ketchup-dispenser line) is folded into Tue + Thu mornings, marked '+CIP' on the Schedule tab. He's OUT of production coverage. Full day-by-day rotation is in the separate 'RI Program' document.")
line("")
line("TRAINING \u2014 EACH SHIFT HAS A CT ON STATION (purple / 'TRAIN'; extra body, not coverage)",bold=True,size=11,fill=GRPFILL,color="FFFFFF")
line("\u2022 Camila (kitchen, ~14 of 25 hrs) and Astrid (drive-thru, 18 of 25 hrs) each shadow a certified trainer on the same station/time, weekdays only (no Sat/Sun). Celeste gets her 3-hr DT training too. Trainees are EXTRA hands; the station is independently covered.")
line("\u2022 Kept the training program running even though trainees aren't on the performer list \u2014 it doesn't take coverage hours from performers. Say the word if you'd rather pause it this week.")
line("")
line("RULES APPLIED",bold=True,size=11,fill=GRPFILL,color="FFFFFF")
line("\u2022 Everyone on at 8:00a at the earliest (no shift before 8a). \u2022 3-hr minimum shift, one continuous shift per person per day. \u2022 No position worked more than 5 hrs (rotate stations); exempt anchors: Avery, Quinn, Tessa, Emil, Felix. \u2022 No close-then-open. \u2022 Registers capped at 2 at once, boards at 3.")
line("\u2022 EZRA = evenings 5:00p\u2013close, open Sunday (per your confirmation; HotSchedules still shows the opposite 12a\u20135p \u2014 worth fixing in HS). \u2022 PALOMA off Tue 7/7 (vacation 7/7\u20137/18). \u2022 Silas & Kofi run front counter solo. \u2022 Tessa front/drive-thru only. \u2022 Paloma drive-thru only.")
line("\u2022 A SECOND register runs at lunch on Wed\u2013Sun (helps speed of service and gives the front-counter performers hours). Say the word and I'll drop it to one register to trim labor.")
line("")
line("COVERAGE & LABOR",bold=True,size=11,fill=GRPFILL,color="FFFFFF")
line("\u2022 ZERO coverage gaps across all seven days. Every required position-hour is staffed.")
line("\u2022 Total ~803 hrs = production coverage at demand + Miles's 20-hr RI + trainee shadow hours. No production padding beyond demand.")
line("")
line("NOT SCHEDULED THIS WEEK",bold=True,size=11,fill=GRPFILL,color="FFFFFF")
line("\u2022 Grady Holt, Lachlan Frost, Maren Kovac, Desmond Ruiz, Callum Voss, Estelle Fontaine, Beau Hendrix (departed / vacation / no current positions).")
line("\u2022 Managers (Manager A, Sama, Manager B, Manager C, Manager D) run the separate MOD line, not the crew grid.")

# ============ PER-DAY POSITION COVERAGE TIMELINES ============
BARFILL={"K":PatternFill("solid",fgColor="F6C26B"),"D":PatternFill("solid",fgColor="8BB8E8"),
         "F":PatternFill("solid",fgColor="9CC79C"),"I":PatternFill("solid",fgColor="C5A3D6")}
GAPFILL=PatternFill("solid",fgColor="E8473F")
SLOTS=[8+0.5*i for i in range(34)]   # 8:00a .. 12:30a ; slot covers [t, t+0.5)
DAYTAB={"Wed":"Wed 7-1","Thu":"Thu 7-2","Fri":"Fri 7-3","Sat":"Sat 7-4","Sun":"Sun 7-5","Mon":"Mon 7-6","Tue":"Tue 7-7"}
TLPOS=["RI","Grill","Board1","Board2","Board3","Toast","DTExp","DTOrd","DTOI","Cash"]
def first(n): return NAMES[n].split()[0]
def req_conc(day,pos,mid): return sum(1 for (s,e) in DEMAND[day].get(pos,[]) if s<=mid<e)
def segs_for(day,pos): return sorted([(s,e,first(n),is_train(n,day,pos,s,e)) for n in assign for (d,p,s,e) in assign[n] if d==day and p==pos])
def lanes_of(segs):
    lanes=[]
    for seg in segs:
        s=seg[0]
        for ln in lanes:
            if ln[-1][1]<=s+1e-9: ln.append(seg); break
        else: lanes.append([seg])
    return lanes
def merge_runs(ws,row,c0,n):
    j=0
    while j<n:
        cell=ws.cell(row,c0+j); val=cell.value; red=(cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb=="00E8473F")
        k=j+1
        while k<n:
            nx=ws.cell(row,c0+k); nred=(nx.fill and nx.fill.fgColor and nx.fill.fgColor.rgb=="00E8473F")
            if (nx.value==val) and (nred==red) and (val is not None or red): k+=1
            else: break
        if k-j>1 and (val is not None or red):
            ws.merge_cells(start_row=row,start_column=c0+j,end_row=row,end_column=c0+k-1)
        j=k
def build_timeline(wb):
    for day in DAYS:
        ws=wb.create_sheet(DAYTAB[day])
        ws.cell(1,1,f"{DATES[day]}  -  POSITION COVERAGE TIMELINE").font=ARIAL(bold=True,size=13,color="FFFFFF")
        ws.cell(1,1).fill=HEADFILL
        ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=2+len(SLOTS))
        ws.cell(2,1,"Each bar = who covers that position over time.  RED = required but uncovered.  PURPLE = trainee shadowing a CT (NOT counted as coverage).  K=Kitchen D=DriveThru F=Front").font=ARIAL(italic=True,size=9,color="555555")
        ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=2+len(SLOTS))
        ws.cell(3,1,"Position").font=ARIAL(bold=True,size=9); ws.cell(3,1).fill=GRPFILL2
        for i,t in enumerate(SLOTS):
            c=ws.cell(3,2+i)
            if abs(t-int(t))<1e-9: c.value=fmt(t).replace(":00",""); c.font=ARIAL(bold=True,size=8)
            c.alignment=Alignment(horizontal="left"); c.fill=GRPFILL2
        row=4
        for pos in TLPOS:
            if not DEMAND[day].get(pos): continue
            peakR=max((req_conc(day,pos,t+0.25) for t in SLOTS),default=0)
            lanes=lanes_of(segs_for(day,pos))
            nrows=max(peakR,len(lanes),1)
            for r in range(nrows):
                lane=lanes[r] if r<len(lanes) else []
                lbl=DISP[pos]+("" if r==0 else f" ({r+1})")
                lc=ws.cell(row,1,lbl); lc.font=ARIAL(size=9,bold=(r==0)); lc.fill=FILL[STN[pos]]; lc.border=BORDER
                lc.alignment=Alignment(vertical="center")
                for i,t in enumerate(SLOTS):
                    mid=t+0.25; cell=ws.cell(row,2+i); cell.border=BORDER
                    seg=next((x for x in lane if x[0]<=mid<x[1]),None)
                    rq=req_conc(day,pos,mid)
                    covered_now=sum(1 for ln in lanes if any((x[0]<=mid<x[1] and not x[3]) for x in ln))
                    if seg:
                        cell.value=seg[2]; cell.fill=TRAINFILL if seg[3] else BARFILL[STN[pos]]
                        cell.font=ARIAL(size=8,italic=bool(seg[3]))
                        cell.alignment=Alignment(horizontal="center",vertical="center")
                    elif covered_now<rq and r<rq:
                        cell.fill=GAPFILL
                merge_runs(ws,row,2,len(SLOTS))
                ws.row_dimensions[row].height=16; row+=1
            row+=0
        ws.column_dimensions["A"].width=13
        for i in range(len(SLOTS)): ws.column_dimensions[get_column_letter(2+i)].width=3.4
        ws.freeze_panes="B4"
build_timeline(wb)
# reorder: Schedule, day tabs, Coverage Check, Hours Summary, Notes
order=["Schedule"]+[DAYTAB[d] for d in DAYS]+["Coverage Check","Hours Summary","Notes & Decisions"]
wb._sheets=[wb[n] for n in order if n in wb.sheetnames]


import os
outdir=os.path.join(os.path.dirname(os.path.abspath(__file__)),"sample_output")
os.makedirs(outdir,exist_ok=True)
out=os.path.join(outdir,"Crew_Schedule_SAMPLE.xlsx")
wb.save(out); print("saved",out)
